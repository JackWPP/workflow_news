"""Import patents from Excel into the Patent table.

Usage:
    python scripts/import_patents.py [--excel PATH]

Reads from 全部专利-纠错后.xlsx (three sheets), merges data by patent_number,
categorizes by keyword rules, and upserts into the patents table.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import session_scope
from app.models import Patent

_PATENT_CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("静电纺丝", ["静电纺", "电纺", "纳米纤维", "纺丝", "原丝", "纺纱"]),
    ("3D打印", ["3D打印", "3d打印", "增材制造", "3D复印"]),
    ("磁流体", ["磁流体", "磁性流体", "磁性液体", "磁悬浮"]),
    ("轮胎", ["轮胎", "胎面", "胎体"]),
    ("传热", ["传热", "换热", "冷却", "散热", "相变", "蓄能"]),
    ("模具硫化", ["模具", "硫化", "模压", "合模"]),
    ("注塑", ["注塑", "注射成型", "注压"]),
    ("挤出", ["挤出", "挤出机", "螺杆", "克拉管"]),
    ("压塑", ["压塑", "压缩成型", "压制", "热压"]),
    ("复合材料", ["复合材料", "纤维增强", "叠层", "复合"]),
    ("纳米", ["纳米", "碳纳米管", "石墨烯"]),
    ("航空航天", ["航空", "航天", "发动机", "无人机"]),
    ("口罩过滤", ["口罩", "滤膜", "过滤", "无纺布"]),
    ("软体机器人", ["软体机器人", "机器人", "人工智能"]),
    ("拉伸成型", ["拉伸", "双向拉伸", "吹塑"]),
    ("燃料电池", ["燃料电池", "氢能", "储氢"]),
    ("微流控", ["微流控", "微流道", "芯片"]),
    ("橡胶", ["橡胶", "弹性体"]),
    ("激光", ["激光", "激光制造", "激光石墨化"]),
]


def classify_patent(name: str) -> str:
    for category, keywords in _PATENT_CATEGORY_RULES:
        for kw in keywords:
            if kw in name:
                return category
    return "其他"


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        s = str(int(value))
        if len(s) == 8:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def read_excel(excel_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    patents: dict[str, dict] = {}

    # Sheet: 总表 (columns: 序号, 名称, 专利号, 授权日, 发明人)
    if "总表" in wb.sheetnames:
        ws = wb["总表"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            pn = str(row[2]).strip() if row[2] else None
            if not pn:
                continue
            patents[pn] = {
                "name": str(row[1]).strip(),
                "patent_number": pn,
                "grant_date": _parse_date(row[3]),
                "inventors": str(row[4]).strip() if row[4] else "",
                "publication_number": None,
                "application_date": None,
            }

    # Sheet: 第一发明人 (columns: 序号, 名称, 专利号, 授权日, 发明人, 公开号, 申请日期)
    if "第一发明人" in wb.sheetnames:
        ws = wb["第一发明人"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            pn = str(row[2]).strip() if row[2] else None
            if not pn:
                continue
            if pn in patents:
                if row[5]:
                    patents[pn]["publication_number"] = str(row[5]).strip()
                if row[6]:
                    patents[pn]["application_date"] = _parse_date(row[6])
            else:
                patents[pn] = {
                    "name": str(row[1]).strip(),
                    "patent_number": pn,
                    "grant_date": _parse_date(row[3]),
                    "inventors": str(row[4]).strip() if row[4] else "",
                    "publication_number": str(row[5]).strip() if row[5] else None,
                    "application_date": _parse_date(row[6]),
                }

    # Sheet: 非第一发明人 (columns: 序号, 名称, 专利号, 授权日期, 发明人, 公开号, 申请日期)
    if "非第一发明人" in wb.sheetnames:
        ws = wb["非第一发明人"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            pn = str(row[2]).strip() if row[2] else None
            if not pn:
                continue
            if pn in patents:
                if row[5] and not patents[pn].get("publication_number"):
                    patents[pn]["publication_number"] = str(row[5]).strip()
                if row[6] and not patents[pn].get("application_date"):
                    patents[pn]["application_date"] = _parse_date(row[6])
            else:
                patents[pn] = {
                    "name": str(row[1]).strip(),
                    "patent_number": pn,
                    "grant_date": _parse_date(row[3]),
                    "inventors": str(row[4]).strip() if row[4] else "",
                    "publication_number": str(row[5]).strip() if row[5] else None,
                    "application_date": _parse_date(row[6]),
                }

    wb.close()

    # Categorize
    result = []
    for p in patents.values():
        p["category"] = classify_patent(p["name"])
        result.append(p)
    return result


def import_patents(patents: list[dict]) -> tuple[int, int]:
    inserted = 0
    updated = 0
    with session_scope() as session:
        for p in patents:
            existing = session.query(Patent).filter_by(patent_number=p["patent_number"]).first()
            if existing:
                existing.name = p["name"]
                existing.grant_date = p["grant_date"]
                existing.inventors = p["inventors"]
                existing.category = p["category"]
                if p.get("publication_number"):
                    existing.publication_number = p["publication_number"]
                if p.get("application_date"):
                    existing.application_date = p["application_date"]
                updated += 1
            else:
                session.add(Patent(
                    name=p["name"],
                    patent_number=p["patent_number"],
                    grant_date=p["grant_date"],
                    inventors=p["inventors"],
                    category=p["category"],
                    publication_number=p.get("publication_number"),
                    application_date=p.get("application_date"),
                ))
                inserted += 1
    return inserted, updated


def main():
    parser = argparse.ArgumentParser(description="Import patents from Excel")
    parser.add_argument("--excel", default="全部专利-纠错后.xlsx", help="Path to Excel file")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"Reading patents from {excel_path}...")
    patents = read_excel(str(excel_path))
    print(f"Found {len(patents)} unique patents.")

    # Category distribution
    from collections import Counter
    dist = Counter(p["category"] for p in patents)
    print("\nCategory distribution:")
    for cat, count in dist.most_common():
        print(f"  {cat}: {count}")

    print(f"\nImporting into database...")
    inserted, updated = import_patents(patents)
    print(f"Done. Inserted: {inserted}, Updated: {updated}")


if __name__ == "__main__":
    main()
